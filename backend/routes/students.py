from fastapi import APIRouter, UploadFile, File, HTTPException
import tempfile

from services.excel_service import parse_student_excel
from services.user_service import create_temp_student
from services.email_templates import password_setup_email
from services.email_service import send_email
from db.collections import users_collection
from core.config import Settings
from openpyxl import load_workbook
import shutil
import os
import tempfile, shutil, os, secrets
from datetime import datetime , timedelta
router = APIRouter(prefix="/students", tags=["Students"])

FRONTEND_URL = Settings.frontend_URL 
def generate_reset_token():
    return secrets.token_urlsafe(32)


def build_password_email(name: str, token: str):
    link = f"{FRONTEND_URL}/set-password?token={token}"

    return f"""
    Hi {name},

    Your Smart Student Hub student account has been created.

    Please click the link below to set your password:

    {link}

    This link will expire in 24 hours.

    Regards,
    Smart Student Hub
    """


@router.post("/upload")
async def upload_students(file: UploadFile = File(...)):

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files allowed")

    try:
        suffix = os.path.splitext(file.filename)[1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            shutil.copyfileobj(file.file, tmp)
            temp_file_path = tmp.name

        wb = load_workbook(temp_file_path)
        ws = wb.active

        created = 0
        skipped = 0
        updated = 0
        errors = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, email, regno, dept = row

            # Validate
            if not email or not name:
                errors.append(f"Missing required fields for row: {row}")
                continue

            existing = await users_collection.find_one({"email": email})

            reset_token = generate_reset_token()
            expiry = datetime.utcnow() + timedelta(hours=24)

            if existing:
                # Optional: update details
                await users_collection.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {
                        "name": name,
                        "register_no": regno,
                        "department": dept,
                        "reset_token": reset_token,
                        "reset_token_exp": expiry
                    }}
                )
                updated += 1
            else:
                await users_collection.insert_one({
                    "name": name,
                    "email": email,
                    "register_no": regno,
                    "department": dept,
                    "role": "student",
                    "password": None,
                    "is_active": False,
                    "reset_token": reset_token,
                    "reset_token_exp": expiry,
                    "created_at": datetime.utcnow()
                })
                created += 1

            # Send email
            await send_email(
                to_email=email,
                subject="Your Smart Student Hub Account",
                html=build_password_email(name, reset_token)
            )


        os.remove(temp_file_path)

        return {
            "message": "Processed successfully",
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors
        }

    except Exception as e:
        raise HTTPException(500, f"Excel read failed: {str(e)}")